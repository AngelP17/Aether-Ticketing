"""One-shot bootstrap: create the legacy tickets schema, load 500 rows from
tickets.xlsx, and align with the Aether migration. Run with the venv Python
from the repo root.
"""
from __future__ import annotations

import os
from datetime import date, datetime

import openpyxl
import psycopg2

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise SystemExit("DATABASE_URL env var is required")

LEGACY_TICKETS_SCHEMA = """
CREATE TABLE IF NOT EXISTS tickets (
    id SERIAL PRIMARY KEY,
    ticket_id VARCHAR(20) UNIQUE NOT NULL,
    title TEXT NOT NULL,
    status VARCHAR(50) DEFAULT 'Open',
    priority VARCHAR(50) DEFAULT 'Low',
    request_type VARCHAR(100),
    staff_assigned VARCHAR(100),
    requester VARCHAR(100),
    date_opened DATE DEFAULT CURRENT_DATE,
    description TEXT,
    resolution_notes TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
"""

EXTRA_LEGACY_TABLES = [
    """
    CREATE TABLE IF NOT EXISTS categories (
        id SERIAL PRIMARY KEY,
        name VARCHAR(100) UNIQUE NOT NULL,
        color VARCHAR(7) DEFAULT '#6366f1',
        icon VARCHAR(50) DEFAULT 'fa-tag',
        is_custom BOOLEAN DEFAULT FALSE,
        is_active BOOLEAN DEFAULT TRUE,
        sort_order INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS assignees (
        id SERIAL PRIMARY KEY,
        name VARCHAR(100) UNIQUE NOT NULL,
        email VARCHAR(255),
        team VARCHAR(100),
        is_active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS labels (
        id SERIAL PRIMARY KEY,
        name VARCHAR(50) UNIQUE NOT NULL,
        color VARCHAR(7) DEFAULT '#6366f1',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS ticket_labels (
        ticket_id VARCHAR(20) REFERENCES tickets(ticket_id) ON DELETE CASCADE,
        label_id INTEGER REFERENCES labels(id) ON DELETE CASCADE,
        PRIMARY KEY (ticket_id, label_id)
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS ticket_comments (
        id SERIAL PRIMARY KEY,
        ticket_id VARCHAR(20) REFERENCES tickets(ticket_id) ON DELETE CASCADE,
        author VARCHAR(100) NOT NULL,
        body TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS ticket_attachments (
        id SERIAL PRIMARY KEY,
        ticket_id VARCHAR(20) REFERENCES tickets(ticket_id) ON DELETE CASCADE,
        filename VARCHAR(255) NOT NULL,
        content_type VARCHAR(100),
        size_bytes INTEGER,
        storage_path TEXT,
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """,
]


def _coerce_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def main() -> None:
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(LEGACY_TICKETS_SCHEMA)
    for stmt in EXTRA_LEGACY_TABLES:
        cur.execute(stmt)

    cur.execute("SELECT COUNT(*) FROM tickets")
    existing = cur.fetchone()[0]
    if existing >= 100:
        print(f"Tickets table already populated ({existing} rows). Skipping Excel import.")
    else:
        wb = openpyxl.load_workbook("tickets.xlsx", data_only=True, read_only=True)
        ws = wb["IT Service Tickets"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            raise SystemExit("tickets.xlsx is empty")

        header_row = next(openpyxl.load_workbook("tickets.xlsx", data_only=True, read_only=True)["IT Service Tickets"].iter_rows(min_row=1, max_row=1, values_only=True))
        idx = {str(name).strip(): i for i, name in enumerate(header_row) if name}

        title_col = idx.get("Ticket Title", idx.get("Title", 1))
        status_col = idx.get("Status", 2)
        priority_col = idx.get("Priority", 3)
        request_type_col = idx.get("Request Type", 4)
        staff_col = idx.get("Staff Assigned", 5)
        requester_col = idx.get("Requester", 6)
        date_col = idx.get("Date", idx.get("Date Opened", 7))
        description_col = idx.get("Description", 8)
        resolution_col = idx.get("Resolution Notes", 9)

        inserted = 0
        for row_idx, row in enumerate(rows, start=2):
            if row is None or all(cell is None for cell in row):
                continue
            title = row[title_col] if title_col < len(row) else None
            if not title:
                continue
            ticket_id = f"IT-2025{row_idx - 1:04d}"
            status = str(row[status_col] or "Open").strip() if status_col < len(row) else "Open"
            priority = str(row[priority_col] or "Low").strip() if priority_col < len(row) else "Low"
            request_type = str(row[request_type_col] or "").strip() if request_type_col < len(row) else ""
            staff = str(row[staff_col] or "").strip() if staff_col < len(row) else ""
            requester = str(row[requester_col] or "").strip() if requester_col < len(row) else ""
            date_opened = _coerce_date(row[date_col]) if date_col < len(row) else None
            description = str(row[description_col] or "").strip() if description_col < len(row) else ""
            resolution = str(row[resolution_col] or "").strip() if resolution_col < len(row) else ""

            cur.execute(
                """
                INSERT INTO tickets (
                    ticket_id, title, status, priority, request_type,
                    staff_assigned, requester, date_opened, description, resolution_notes
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticket_id) DO NOTHING
                """,
                (
                    ticket_id,
                    str(title).strip(),
                    status or "Open",
                    priority or "Low",
                    request_type or None,
                    staff or None,
                    requester or None,
                    date_opened,
                    description or None,
                    resolution or None,
                ),
            )
            inserted += 1
        conn.commit()
        print(f"Imported {inserted} tickets from tickets.xlsx")

    cur.execute("SELECT COUNT(*) FROM tickets")
    final = cur.fetchone()[0]
    print(f"Tickets table now has {final} rows")

    cur.execute("SELECT status, COUNT(*) FROM tickets GROUP BY status ORDER BY COUNT(*) DESC")
    for status, count in cur.fetchall():
        print(f"  {status}: {count}")
    cur.execute("SELECT priority, COUNT(*) FROM tickets GROUP BY priority ORDER BY COUNT(*) DESC")
    for priority, count in cur.fetchall():
        print(f"  {priority}: {count}")

    conn.close()


if __name__ == "__main__":
    main()
