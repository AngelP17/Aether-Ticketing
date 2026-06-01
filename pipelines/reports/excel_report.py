"""
Excel Report Generator: Produces the 5-tab styled operational workbook.
Tab 1: Executive Summary | Tab 2: Operational Queue | Tab 3: Incident Clusters
Tab 4: Decision Intelligence | Tab 5: Audit Extract
"""

from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="404040"),
    right=Side(style="thin", color="404040"),
    top=Side(style="thin", color="404040"),
    bottom=Side(style="thin", color="404040"),
)

PRIORITY_FILLS = {
    "Critical": PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid"),
    "High": PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid"),
    "Medium": PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid"),
    "Low": PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid"),
}

STATUS_FILLS = {
    "Open": PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid"),
    "In Progress": PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid"),
    "Waiting for Info": PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid"),
    "Resolved": PatternFill(start_color="ccfbf1", end_color="ccfbf1", fill_type="solid"),
    "Closed": PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid"),
}


def generate_workbook(
    report_type: str = "operational",
    *,
    tickets: list[dict[str, Any]] | None = None,
    incidents: list[dict[str, Any]] | None = None,
) -> Workbook:
    tickets = tickets or []
    incidents = incidents or []

    wb = Workbook()

    _build_executive_summary(wb, tickets=tickets, incidents=incidents)
    _build_operational_queue(wb, tickets=tickets)
    _build_incident_clusters(wb, incidents=incidents)
    _build_decision_intelligence(wb, tickets=tickets)
    _build_audit_extract(wb, tickets=tickets)

    return wb


def _build_executive_summary(wb: Workbook, *, tickets: list[dict[str, Any]], incidents: list[dict[str, Any]]) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Executive Summary"

    ws.cell(1, 1, "Aether — Executive Summary").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Generated {datetime.now(UTC).isoformat()}").font = Font(color="666666", size=10)

    _apply_header_style(ws, row=4, headers=["KPI", "Value"])

    open_tickets = [ticket for ticket in tickets if ticket["status"] not in {"Resolved", "Closed"}]
    critical = [ticket for ticket in open_tickets if ticket["priority_raw"] == "Critical"]
    sla_risk = [ticket for ticket in open_tickets if (ticket.get("sla_risk") or 0) >= 75]
    kpis = [
        ("Total Open Tickets", len(open_tickets)),
        ("Critical Queue", len(critical)),
        ("SLA Breach Risk", len(sla_risk)),
        ("Incident Clusters", len(incidents)),
    ]
    for i, (kpi, formula) in enumerate(kpis, start=5):
        ws.cell(i, 1, kpi)
        ws.cell(i, 2, formula)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def _build_operational_queue(wb: Workbook, *, tickets: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Operational Queue")

    headers = [
        "Ticket ID",
        "Title",
        "Status",
        "Priority",
        "Category",
        "Assignee",
        "Site",
        "Days Open",
        "Priority Score",
        "Root Cause",
        "SLA Risk",
        "Confidence",
        "Recommendation",
    ]
    _apply_header_style(ws, row=1, headers=headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_index, ticket in enumerate(tickets, start=2):
        values = [
            ticket["ticket_id"],
            ticket["title"],
            ticket["status"],
            ticket["priority_raw"],
            ticket.get("category") or "",
            ticket.get("assignee") or "",
            ticket.get("site") or "",
            ticket["days_open"],
            ticket.get("priority_score"),
            ticket.get("root_cause_hypothesis") or "",
            ticket.get("sla_risk") or 0,
            ticket.get("confidence_score") or 0,
            ticket.get("recommendation") or "",
        ]
        _append_row(ws, row_index, values)
        if ticket["priority_raw"] in PRIORITY_FILLS:
            ws.cell(row_index, 4).fill = PRIORITY_FILLS[ticket["priority_raw"]]
        if ticket["status"] in STATUS_FILLS:
            ws.cell(row_index, 3).fill = STATUS_FILLS[ticket["status"]]
    _autosize(ws, len(headers))


def _build_incident_clusters(wb: Workbook, *, incidents: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Incident Clusters")

    headers = [
        "Incident ID",
        "Title",
        "Status",
        "Root Cause",
        "Ticket Count",
        "Confidence",
        "Impact Score",
    ]
    _apply_header_style(ws, row=1, headers=headers)
    for row_index, incident in enumerate(incidents, start=2):
        _append_row(
            ws,
            row_index,
            [
                incident["id"],
                incident["title"],
                incident["status"],
                incident["root_cause_hypothesis"],
                incident["ticket_count"],
                incident["confidence"],
                incident["business_impact_score"],
            ],
        )
    _autosize(ws, len(headers))


def _build_decision_intelligence(wb: Workbook, *, tickets: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Decision Intelligence")

    headers = [
        "Ticket ID",
        "Priority Score",
        "Severity",
        "Urgency",
        "Business Impact",
        "SLA Risk",
        "Recurrence",
        "Actionability",
        "Uncertainty",
        "Root Cause",
        "Confidence",
        "Top Recommendation",
    ]
    _apply_header_style(ws, row=1, headers=headers)
    for row_index, ticket in enumerate(tickets, start=2):
        _append_row(
            ws,
            row_index,
            [
                ticket["ticket_id"],
                ticket.get("priority_score") or 0,
                ticket.get("severity_score") or 0,
                ticket.get("urgency_score") or 0,
                ticket.get("business_impact_score") or 0,
                ticket.get("sla_risk") or 0,
                ticket.get("recurrence_score") or 0,
                ticket.get("actionability_score") or 0,
                ticket.get("uncertainty_penalty") or 0,
                ticket.get("root_cause_hypothesis") or "",
                ticket.get("confidence_score") or 0,
                ticket.get("recommendation") or "",
            ],
        )
    _autosize(ws, len(headers))


def _build_audit_extract(wb: Workbook, *, tickets: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Audit Extract")

    headers = ["Ticket ID", "Event Type", "Actor", "Timestamp", "Details", "Feedback Status"]
    _apply_header_style(ws, row=1, headers=headers)
    for row_index, ticket in enumerate(tickets, start=2):
        _append_row(
            ws,
            row_index,
            [
                ticket["ticket_id"],
                "decision_generated",
                "aether-api",
                datetime.now(UTC).isoformat(),
                ticket.get("root_cause_hypothesis") or "",
                "proposed",
            ],
        )
    _autosize(ws, len(headers))


def _apply_header_style(ws: Any, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _append_row(ws: Any, row_index: int, values: list[object]) -> None:
    for column, value in enumerate(values, start=1):
        cell = ws.cell(row=row_index, column=column, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws: Any, column_count: int) -> None:
    for column in range(1, column_count + 1):
        letter = get_column_letter(column)
        width = max(
            len(str(ws.cell(row=row, column=column).value or ""))
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 42)
