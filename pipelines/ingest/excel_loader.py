"""
Excel Loader: Reads tickets.xlsx and yields normalized ticket rows.
Compatible with the existing IT Service Tickets sheet format.
"""

from openpyxl import load_workbook
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Iterator, cast
import hashlib


EXCEL_FILE = "tickets.xlsx"
SHEET_NAME = "IT Service Tickets"


@dataclass
class NormalizedTicket:
    external_ticket_id: str
    title: str
    status: str
    priority_raw: str
    request_type: str
    staff_assigned: str
    requester: str
    date_opened: datetime
    description: str
    resolution_notes: str
    source_hash: str


def compute_row_hash(row_data: dict[str, Any]) -> str:
    content = f"{row_data.get('title', '')}|{row_data.get('status', '')}|{row_data.get('description', '')}"
    return hashlib.sha256(content.encode()).hexdigest()[:16]


def load_excel_rows(
    excel_file: str = EXCEL_FILE, sheet_name: str = SHEET_NAME
) -> Iterator[NormalizedTicket]:
    """Load and normalize all ticket rows from the Excel file."""
    try:
        wb = load_workbook(excel_file, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]

    for row_num in range(2, 10000):
        title_cell = ws[f"B{row_num}"].value
        if title_cell is None:
            break

        raw_data = {
            "title": str(title_cell).strip() if title_cell else "",
            "status": str(ws[f"C{row_num}"].value or "Open").strip(),
            "priority": str(ws[f"D{row_num}"].value or "Low").strip(),
            "request_type": str(ws[f"E{row_num}"].value or "").strip(),
            "staff_assigned": str(ws[f"F{row_num}"].value or "").strip(),
            "requester": str(ws[f"G{row_num}"].value or "").strip(),
            "date_opened": ws[f"H{row_num}"].value,
            "description": str(ws[f"J{row_num}"].value or "").strip(),
            "resolution_notes": str(ws[f"K{row_num}"].value or "").strip(),
        }

        ticket_id = str(ws[f"A{row_num}"].value or f"IT-2025{row_num - 1:04d}").strip()
        if not ticket_id.startswith("IT-"):
            ticket_id = f"IT-2025{row_num - 1:04d}"

        date_opened = _parse_date(raw_data["date_opened"])

        yield NormalizedTicket(
            external_ticket_id=ticket_id,
            title=raw_data["title"],
            status=raw_data["status"],
            priority_raw=raw_data["priority"],
            request_type=raw_data["request_type"],
            staff_assigned=raw_data["staff_assigned"],
            requester=raw_data["requester"],
            date_opened=date_opened,
            description=raw_data["description"],
            resolution_notes=raw_data["resolution_notes"],
            source_hash=compute_row_hash(raw_data),
        )

    wb.close()


def _parse_date(value: object) -> datetime:
    if value is None:
        return datetime.now()
    if hasattr(value, "date"):
        return datetime.combine(value.date(), datetime.min.time())
    if hasattr(value, "strftime"):
        return cast(datetime, value)
    try:
        return datetime.strptime(str(value), "%Y-%m-%d")
    except Exception:
        return datetime.now()
